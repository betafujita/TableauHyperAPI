import xml.etree.ElementTree as ET
import json
import zipfile
import os
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

def extract_twb_from_twbx(twbx_file_path):
    """TWBXファイルからTWBファイル（設定情報が含まれるXML）を抽出"""
    try:
        with zipfile.ZipFile(twbx_file_path, 'r') as z:
            twb_files = [f for f in z.namelist() if f.endswith('.twb')]
            if not twb_files:
                print("エラー: TWBXファイル内にTWBファイルが見つかりません")
                return None
            
            twb_file = twb_files[0]
            print(f"TWBX内から抽出したTWBファイル: {twb_file}")
            return z.read(twb_file)
    except FileNotFoundError:
        print(f"エラー: ファイル {twbx_file_path} が見つかりません")
        return None
    except Exception as e:
        print(f"エラー: TWBXファイルの解凍に失敗しました - {e}")
        return None

def extract_tableau_info(twb_file_content):
    """TWBファイルからTableau設定情報を抽出"""
    try:
        root = ET.fromstring(twb_file_content)
        print("ファイルの解析に成功しました")
    except Exception as e:
        print(f"エラー: ファイルの解析に失敗しました - {e}")
        return None
    
    result = {
        "データソース": [],
        "計算フィールド": [],
        "リレーション": []
    }
    
    try:
        for datasource in root.findall('.//datasource'):
            ds_info = {
                "名前": datasource.get('name', '不明'),
                "タイプ": datasource.get('type', '不明'),
                "カラム": []
            }
            
            for column in datasource.findall('.//column'):
                col_info = {
                    "名前": column.get('name', '不明'),
                    "データ型": column.get('datatype', '不明'),
                    "表示名": column.get('caption', column.get('name', '不明'))
                }
                
                calc = column.find('./calculation')
                if calc is not None:
                    formula = calc.get('formula', '')
                    col_info["計算式"] = formula
                    
                    result["計算フィールド"].append({
                        "名前": col_info["名前"],
                        "表示名": col_info["表示名"],
                        "データソース": ds_info["名前"],
                        "計算式": formula
                    })
                
                ds_info["カラム"].append(col_info)
            
            for relation in datasource.findall('.//relation'):
                rel_info = {
                    "タイプ": relation.get('type', '不明'),
                    "名前": relation.get('name', '不明')
                }
                
                if rel_info["タイプ"] == 'join':
                    rel_info["結合タイプ"] = relation.get('join', '不明')
                    
                clause = relation.find('./clause')
                if clause is not None:
                    rel_info["結合条件"] = ET.tostring(clause, encoding='unicode')
                
                result["リレーション"].append(rel_info)
            
            result["データソース"].append(ds_info)
        
        print(f"データソース数: {len(result['データソース'])}")
        print(f"計算フィールド数: {len(result['計算フィールド'])}")
        print(f"リレーション数: {len(result['リレーション'])}")
    except Exception as e:
        print(f"エラー: データ抽出中に問題が発生しました - {e}")
    
    return result

def style_header_cell(cell):
    """ヘッダーセルのスタイルを設定"""
    cell.font = Font(bold=True, color="FFFFFF", size=11)
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def style_data_cell(cell, wrap=False):
    """データセルのスタイルを設定"""
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

def create_manual_sheet(wb):
    """マニュアルシートを作成"""
    ws = wb.create_sheet("マニュアル", 0)
    
    # タイトル
    ws['A1'] = "Tableauワークブック情報抽出ツール - マニュアル"
    ws['A1'].font = Font(bold=True, size=16, color="4472C4")
    ws.merge_cells('A1:D1')
    
    row = 3
    
    # 概要セクション
    ws[f'A{row}'] = "■ 概要"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="4472C4")
    row += 1
    
    ws[f'A{row}'] = "このツールは、Tableauワークブック(.twbx)から設定情報を抽出し、Excelブックとして出力します。"
    ws.merge_cells(f'A{row}:D{row}')
    row += 2
    
    # 各シートの説明
    ws[f'A{row}'] = "■ 各シートの説明"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="4472C4")
    row += 1
    
    sheet_descriptions = [
        ("マニュアル", "このシート。ツールの使い方と各シートの説明を記載しています。"),
        ("サマリー", "抽出されたデータソース、計算フィールド、リレーションの件数を表示します。"),
        ("データソース一覧", "Tableauワークブックで使用されているデータソースの一覧を表示します。"),
        ("カラム一覧", "各データソースに含まれるカラム（フィールド）の詳細情報を表示します。"),
        ("計算フィールド一覧", "Tableau内で作成された計算フィールドと、その計算式を表示します。"),
        ("リレーション一覧", "データソース間のリレーション（結合）情報を表示します。"),
    ]
    
    ws[f'A{row}'] = "シート名"
    ws[f'B{row}'] = "説明"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1
    
    for sheet_name, description in sheet_descriptions:
        ws[f'A{row}'] = sheet_name
        ws[f'B{row}'] = description
        ws[f'A{row}'].alignment = Alignment(vertical="top")
        ws[f'B{row}'].alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    
    row += 1
    
    # 使い方セクション
    ws[f'A{row}'] = "■ コマンドライン使用方法"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="4472C4")
    row += 1
    
    usage_examples = [
        ("基本的な使い方", "python tableau_info_extractor.py sales_dashboard.twbx"),
        ("出力先を指定", "python tableau_info_extractor.py sales_dashboard.twbx -o ./output"),
        ("ヘルプを表示", "python tableau_info_extractor.py -h"),
    ]
    
    ws[f'A{row}'] = "使用例"
    ws[f'B{row}'] = "コマンド"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1
    
    for example, command in usage_examples:
        ws[f'A{row}'] = example
        ws[f'B{row}'] = command
        ws[f'A{row}'].alignment = Alignment(vertical="top")
        ws[f'B{row}'].alignment = Alignment(vertical="top")
        ws[f'B{row}'].font = Font(name='Courier New')
        row += 1
    
    row += 1
    
    # 用語解説
    ws[f'A{row}'] = "■ 用語解説"
    ws[f'A{row}'].font = Font(bold=True, size=14, color="4472C4")
    row += 1
    
    terms = [
        ("TWBX", "Tableauのパッケージド・ワークブックファイル。データソースやビジュアライゼーションを含む圧縮ファイル。"),
        ("データソース", "Tableauが接続するデータの出所（データベース、ファイルなど）。"),
        ("カラム/フィールド", "データソース内の個々のデータ項目。"),
        ("計算フィールド", "Tableau内で作成された計算式を持つフィールド。"),
        ("リレーション", "データソース間の関連付け方法（結合など）。"),
    ]
    
    ws[f'A{row}'] = "用語"
    ws[f'B{row}'] = "説明"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'A{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    ws[f'B{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    row += 1
    
    for term, description in terms:
        ws[f'A{row}'] = term
        ws[f'B{row}'] = description
        ws[f'A{row}'].alignment = Alignment(vertical="top")
        ws[f'B{row}'].alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 80

def create_summary_sheet(wb, result, input_file):
    """サマリーシートを作成"""
    ws = wb.create_sheet("サマリー")
    
    # タイトル
    ws['A1'] = "Tableauワークブック情報 - サマリー"
    ws['A1'].font = Font(bold=True, size=14, color="4472C4")
    ws.merge_cells('A1:B1')
    
    # 基本情報
    row = 3
    ws[f'A{row}'] = "抽出元ファイル"
    ws[f'B{row}'] = os.path.basename(input_file)
    ws[f'A{row}'].font = Font(bold=True)
    row += 1
    
    ws[f'A{row}'] = "抽出日時"
    ws[f'B{row}'] = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
    ws[f'A{row}'].font = Font(bold=True)
    row += 2
    
    # 統計情報
    ws[f'A{row}'] = "項目"
    ws[f'B{row}'] = "件数"
    style_header_cell(ws[f'A{row}'])
    style_header_cell(ws[f'B{row}'])
    row += 1
    
    stats = [
        ("データソース数", len(result["データソース"])),
        ("計算フィールド数", len(result["計算フィールド"])),
        ("リレーション数", len(result["リレーション"])),
    ]
    
    for label, count in stats:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = count
        style_data_cell(ws[f'A{row}'])
        style_data_cell(ws[f'B{row}'])
        row += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15

def create_datasources_sheet(wb, result):
    """データソース一覧シートを作成"""
    ws = wb.create_sheet("データソース一覧")
    
    # ヘッダー
    headers = ["No.", "データソース名", "タイプ", "カラム数"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell)
    
    # データ
    for idx, ds in enumerate(result["データソース"], 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=ds["名前"])
        ws.cell(row=idx+1, column=3, value=ds["タイプ"])
        ws.cell(row=idx+1, column=4, value=len(ds["カラム"]))
        
        for col in range(1, 5):
            style_data_cell(ws.cell(row=idx+1, column=col))
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15

def create_columns_sheet(wb, result):
    """カラム一覧シートを作成"""
    ws = wb.create_sheet("カラム一覧")
    
    # ヘッダー
    headers = ["No.", "データソース", "カラム名", "表示名", "データ型", "計算フィールド"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell)
    
    # データ
    row_num = 2
    for ds in result["データソース"]:
        for col in ds["カラム"]:
            ws.cell(row=row_num, column=1, value=row_num-1)
            ws.cell(row=row_num, column=2, value=ds["名前"])
            ws.cell(row=row_num, column=3, value=col["名前"])
            ws.cell(row=row_num, column=4, value=col["表示名"])
            ws.cell(row=row_num, column=5, value=col["データ型"])
            ws.cell(row=row_num, column=6, value="○" if "計算式" in col else "")
            
            for c in range(1, 7):
                style_data_cell(ws.cell(row=row_num, column=c))
            
            row_num += 1
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15

def create_calculated_fields_sheet(wb, result):
    """計算フィールド一覧シートを作成"""
    ws = wb.create_sheet("計算フィールド一覧")
    
    # ヘッダー
    headers = ["No.", "データソース", "フィールド名", "表示名", "計算式"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell)
    
    # データ
    for idx, calc in enumerate(result["計算フィールド"], 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=calc["データソース"])
        ws.cell(row=idx+1, column=3, value=calc["名前"])
        ws.cell(row=idx+1, column=4, value=calc["表示名"])
        ws.cell(row=idx+1, column=5, value=calc["計算式"])
        
        for col in range(1, 6):
            cell = ws.cell(row=idx+1, column=col)
            style_data_cell(cell, wrap=True if col == 5 else False)
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 35
    ws.column_dimensions['E'].width = 60

def create_relations_sheet(wb, result):
    """リレーション一覧シートを作成"""
    ws = wb.create_sheet("リレーション一覧")
    
    # ヘッダー
    headers = ["No.", "名前", "タイプ", "結合タイプ", "結合条件"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        style_header_cell(cell)
    
    # データ
    for idx, rel in enumerate(result["リレーション"], 1):
        ws.cell(row=idx+1, column=1, value=idx)
        ws.cell(row=idx+1, column=2, value=rel["名前"])
        ws.cell(row=idx+1, column=3, value=rel["タイプ"])
        ws.cell(row=idx+1, column=4, value=rel.get("結合タイプ", ""))
        ws.cell(row=idx+1, column=5, value=rel.get("結合条件", ""))
        
        for col in range(1, 6):
            cell = ws.cell(row=idx+1, column=col)
            style_data_cell(cell, wrap=True if col == 5 else False)
    
    # 列幅を調整
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 60

def save_results_excel(result, input_file_path, output_dir=None):
    """結果をExcelブックとして保存"""
    if output_dir is None:
        output_dir = os.path.join(os.path.dirname(input_file_path) or '.', 'Tableau_Info')
    
    os.makedirs(output_dir, exist_ok=True)
    
    base_filename = os.path.splitext(os.path.basename(input_file_path))[0]
    output_excel_filename = os.path.join(output_dir, f"{base_filename}_info.xlsx")
    
    # Excelブックを作成
    wb = Workbook()
    
    # デフォルトのシートを削除
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # 各シートを作成
    create_manual_sheet(wb)
    create_summary_sheet(wb, result, input_file_path)
    create_datasources_sheet(wb, result)
    create_columns_sheet(wb, result)
    create_calculated_fields_sheet(wb, result)
    create_relations_sheet(wb, result)
    
    # 保存
    wb.save(output_excel_filename)
    
    print(f"\n処理が完了しました")
    print(f"Excelファイルは以下の場所に保存されました:")
    print(f"- {output_excel_filename}")
    
    return output_excel_filename

def main():
    """メイン処理"""
    parser = argparse.ArgumentParser(
        description='Tableauワークブック(.twbx)から設定情報を抽出してExcel形式で出力するツール',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
使用例:
  python tableau_info_extractor.py sales_dashboard.twbx
  python tableau_info_extractor.py sales_dashboard.twbx -o ./output
  python tableau_info_extractor.py sales_dashboard.twbx --output-dir ./output
        '''
    )
    parser.add_argument('input_file', help='Tableauワークブックファイル(.twbx)のパス')
    parser.add_argument('-o', '--output-dir', help='出力ディレクトリのパス（省略時は入力ファイルと同じディレクトリに作成）')
    
    args = parser.parse_args()
    
    if not os.path.exists(args.input_file):
        print(f"エラー: ファイル '{args.input_file}' が見つかりません")
        sys.exit(1)
    
    if not args.input_file.lower().endswith('.twbx'):
        print("警告: 入力ファイルの拡張子が .twbx ではありません")
    
    print(f"Tableauワークブック '{args.input_file}' を読み込んでいます...")
    
    try:
        twb_content = extract_twb_from_twbx(args.input_file)
        
        if twb_content:
            result = extract_tableau_info(twb_content)
            
            if result:
                save_results_excel(result, args.input_file, args.output_dir)
                
                json_result = json.dumps(result, ensure_ascii=False, indent=2)
                print("\n抽出されたデータの一部:")
                print(json_result[:500] + "..." if len(json_result) > 500 else json_result)
            else:
                print("情報の抽出に失敗しました")
                sys.exit(1)
        else:
            print("TWBファイルの抽出に失敗したため、処理を中止します")
            sys.exit(1)
    except Exception as e:
        print(f"予期せぬエラーが発生しました: {e}")
        print("詳細なエラー情報:")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
